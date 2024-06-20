from PyQt5 import QtCore, QtGui, QtWidgets
import os
from openpyxl import Workbook
import sqlite3
from datetime import datetime

timecount = 0

class secondwindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1100, 600)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(0, 0, 1100, 600))
        self.label.setObjectName("label")
        self.label.setPixmap(QtGui.QPixmap("002.jpg"))
        
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(865, 128, 150, 25))
        self.label_2.setObjectName("label_2")
        self.label_2.setStyleSheet("color: white;")
        
        self.lineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit.setGeometry(QtCore.QRect(935, 210, 75, 25))
        self.lineEdit.setObjectName("lineEdit")
        self.lineEdit.setStyleSheet("border: none; color: black; background-color: rgba(0, 0, 0, 0.0);")
        
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(821, 269, 105, 30))
        self.pushButton.setObjectName("pushButton")
        self.pushButton.setStyleSheet("background-color: rgba(0, 0, 0, 0.0);") 
        self.pushButton.clicked.connect(self.open1)  
        
        self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(950, 269, 105, 30))
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_2.setStyleSheet("background-color: rgba(0, 0, 0, 0.0);") 
        self.pushButton_2.clicked.connect(self.open2)  
        
    def open1(self):
        text = self.lineEdit.text()
        file = f"C:\\study\\과제\\excel\\{text}.xlsx"
        con = sqlite3.connect("login_records.db")
        cur = con.cursor() 
        
        cur.execute("INSERT INTO login_records (search_file) VALUES (?)", (text,))
        con.commit()
        con.close()
        
        if os.path.isfile(file): 
            os.startfile(file)
        else:
            self.label_2.setText("파일이 존재하지 않습니다.")
            
    def open2(self):
        text = self.lineEdit.text()
        con = sqlite3.connect("login_records.db")
        cur = con.cursor() 
        
        cur.execute("INSERT INTO login_records (make_file) VALUES (?)", (text,))
        con.commit()
        con.close()
        
        # 새로운 워크북(엑셀 파일) 생성
        wb = Workbook()

        # 첫 번째 시트 생성
        sheet1 = wb.active
        sheet1.title = "1반_1학기_중간"

        # 시트1에 데이터 작성
        sheet1['A1'] = '반'
        sheet1['B1'] = '번호'
        sheet1['C1'] = '이름'
        sheet1['D1'] = '과목'
        sheet1['E1'] = '점수'
        sheet1['F1'] = '등수'
        
        # 1열에 1 설정
        for i in range(2, 12):  # 행 개수에 맞게 반복
            sheet1[f'A{i}'] = '1반'  # 1열의 모든 행에 1 설정
            
        for i in range(1, 11): 
            sheet1[f'B{i+1}'] = str(i) + '번' 

        # 두 번째 시트 생성
        sheet2 = wb.create_sheet(title="1반_1학기_기말")

        # 시트2에 데이터 작성
        sheet2['A1'] = '반'
        sheet2['B1'] = '번호'
        sheet2['C1'] = '이름'
        sheet2['D1'] = '과목'
        sheet2['E1'] = '점수'
        sheet2['F1'] = '등수'
        
        for i in range(2, 12):  # 행 개수에 맞게 반복
            sheet2[f'A{i}'] = '1반'  # 1열의 모든 행에 1 설정
            
        for i in range(1, 11): 
            sheet2[f'B{i+1}'] = str(i) + '번' 
        
        # 세 번째 시트 생성
        sheet3 = wb.create_sheet(title="1반_2학기_중간")

        # 시트3에 데이터 작성
        sheet3['A1'] = '반'
        sheet3['B1'] = '번호'
        sheet3['C1'] = '이름'
        sheet3['D1'] = '과목'
        sheet3['E1'] = '점수'
        sheet3['F1'] = '등수'
        
        for i in range(2, 12):  # 행 개수에 맞게 반복
            sheet3[f'A{i}'] = '1반'  # 1열의 모든 행에 1 설정
            
        for i in range(1, 11): 
            sheet3[f'B{i+1}'] = str(i) + '번' 
        
        # 네 번째 시트 생성
        sheet4 = wb.create_sheet(title="1반_2학기_기말")

        # 시트4에 데이터 작성
        sheet4['A1'] = '반'
        sheet4['B1'] = '번호'
        sheet4['C1'] = '이름'
        sheet4['D1'] = '과목'
        sheet4['E1'] = '점수'
        sheet4['F1'] = '등수'
        
        for i in range(2, 12):  # 행 개수에 맞게 반복
            sheet4[f'A{i}'] = '1반'  # 1열의 모든 행에 1 설정
            
        for i in range(1, 11): 
            sheet4[f'B{i+1}'] = str(i) + '번' 
        
        sheet5 = wb.create_sheet(title = "2반_1학기_중간")
        sheet5['A1'] = '반'
        sheet5['B1'] = '번호'
        sheet5['C1'] = '이름'
        sheet5['D1'] = '과목'
        sheet5['E1'] = '점수'
        sheet5['F1'] = '등수'
        
        for i in range(2, 13): 
            sheet5[f'A{i}'] = '2반'  
            
        for i in range(1, 12): 
            sheet5[f'B{i+1}'] = str(i) + '번' 

        sheet6 = wb.create_sheet(title="2반_1학기_기말")
        sheet6['A1'] = '반'
        sheet6['B1'] = '번호'
        sheet6['C1'] = '이름'
        sheet6['D1'] = '과목'
        sheet6['E1'] = '점수'
        sheet6['F1'] = '등수'
        
        for i in range(2, 13): 
            sheet6[f'A{i}'] = '2반' 
            
        for i in range(1, 12): 
            sheet6[f'B{i+1}'] = str(i) + '번' 
        
        sheet7 = wb.create_sheet(title="2반_2학기_중간")
        sheet7['A1'] = '반'
        sheet7['B1'] = '번호'
        sheet7['C1'] = '이름'
        sheet7['D1'] = '과목'
        sheet7['E1'] = '점수'
        sheet7['F1'] = '등수'
        
        for i in range(2, 13): 
            sheet7[f'A{i}'] = '2반' 
            
        for i in range(1, 12): 
            sheet7[f'B{i+1}'] = str(i) + '번' 
        
        sheet8 = wb.create_sheet(title="2반_2학기_기말")
        sheet8['A1'] = '반'
        sheet8['B1'] = '번호'
        sheet8['C1'] = '이름'
        sheet8['D1'] = '과목'
        sheet8['E1'] = '점수'
        sheet8['F1'] = '등수'
        
        for i in range(2, 13): 
            sheet8[f'A{i}'] = '2반' 
            
        for i in range(1, 12): 
            sheet8[f'B{i+1}'] = str(i) + '번' 
        
        sheet9 = wb.create_sheet(title = "3반_1학기_중간")
        sheet9['A1'] = '반'
        sheet9['B1'] = '번호'
        sheet9['C1'] = '이름'
        sheet9['D1'] = '과목'
        sheet9['E1'] = '점수'
        sheet9['F1'] = '등수'
        
        for i in range(2, 15): 
            sheet9[f'A{i}'] = '3반' 
            
        for i in range(1, 14): 
            sheet9[f'B{i+1}'] = str(i) + '번' 

        sheet10 = wb.create_sheet(title="3반_1학기_기말")
        sheet10['A1'] = '반'
        sheet10['B1'] = '번호'
        sheet10['C1'] = '이름'
        sheet10['D1'] = '과목'
        sheet10['E1'] = '점수'
        sheet10['F1'] = '등수'
        
        for i in range(2, 15): 
            sheet10[f'A{i}'] = '3반' 
            
        for i in range(1, 14): 
            sheet10[f'B{i+1}'] = str(i) + '번' 
        
        sheet11 = wb.create_sheet(title="3반_2학기_중간")
        sheet11['A1'] = '반'
        sheet11['B1'] = '번호'
        sheet11['C1'] = '이름'
        sheet11['D1'] = '과목'
        sheet11['E1'] = '점수'
        sheet11['F1'] = '등수'
        
        for i in range(2, 15): 
            sheet11[f'A{i}'] = '3반' 
            
        for i in range(1, 14): 
            sheet11[f'B{i+1}'] = str(i) + '번' 
        
        sheet12 = wb.create_sheet(title="3반_2학기_기말")
        sheet12['A1'] = '반'
        sheet12['B1'] = '번호'
        sheet12['C1'] = '이름'
        sheet12['D1'] = '과목'
        sheet12['E1'] = '점수'
        sheet12['F1'] = '등수'
        
        for i in range(2, 15): 
            sheet12[f'A{i}'] = '3반' 
        
        for i in range(1, 14): 
            sheet12[f'B{i+1}'] = str(i) + '번' 


        # 엑셀 파일 저장
        file_path = f"C:\\study\\과제\\excel\\{text}.xlsx"
        if os.path.isfile(file_path):
            self.label_2.setText("이미 파일이 있습니다.")
        else:
            wb.save(file_path)
            self.label_2.setText("새파일이 생성되었습니다.")


class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1100, 600)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(0, 0, 1100, 600))
        self.label.setObjectName("label")
        self.label.setPixmap(QtGui.QPixmap("001.jpg"))
        
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(125, 480, 170, 20))
        self.label_2.setObjectName("label_2")
        
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(125, 500, 70, 20))
        self.label_3.setObjectName("label_3")
        
        self.lineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit.setGeometry(QtCore.QRect(281, 340, 288, 26))
        self.lineEdit.setObjectName("lineEdit")
        self.lineEdit.setStyleSheet("border: none; background-color: rgba(0, 0, 0, 0.0);")
        
        self.lineEdit_2 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_2.setGeometry(QtCore.QRect(281, 396, 288, 26))
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.lineEdit_2.setStyleSheet("border: none; background-color: rgba(0, 0, 0, 0.0);")
        
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(545, 475, 100, 30))
        self.pushButton.setObjectName("pushButton")
        self.pushButton.setStyleSheet("background-color: rgba(0, 0, 0, 0.0);")
        
        MainWindow.setCentralWidget(self.centralwidget)
        
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        
        self.pushButton.pressed.connect(self.rejult)
        
        self.incorrect_attempts = 0
        self.timer = QtCore.QTimer()
        self.timer.timeout.connect(self.exit_application)
        self.remaining_seconds = 5  
     
    def rejult(self):
        text = self.lineEdit_2.text()
        name = self.lineEdit.text()
        login_time = datetime.now().strftime("%Y-%m-%d %H:%M")

        if text == '0802' and name == "" :
            self.label_2.setText("아이디를 입력하세요.") 
            
        elif text == "" :
            self.label_2.setText("비밀번호를 입력하세요.") 
            
        elif text == '0802':
            self.window = QtWidgets.QWidget()
            self.ui = secondwindow()
            self.ui.setupUi(self.window)
            self.window.show()
            MainWindow.hide()
            con = sqlite3.connect("login_records.db")
            cur = con.cursor() 
            
            cur.execute('''CREATE TABLE IF NOT EXISTS login_records (
                    id INTEGER PRIMARY KEY,
                    login_id TEXT,
                    login_time TEXT,
                    search_file TEXT,
                    make_file TEXT
                    )''')
            
            cur.execute("INSERT INTO login_records (login_id, login_time) VALUES (?, ?)", (name, login_time))
            con.commit()
            con.close()
            
        else:
            self.label_2.setHidden(False)
            self.label_3.setHidden(False)
            self.incorrect_attempts +=1
            self.label_2.setText("틀렸습니다.")     
            if self.incorrect_attempts < 6:
                self.label_3.setText(f'{self.incorrect_attempts}/5')
                self.label_3.setHidden(False)
                if self.incorrect_attempts == 5:
                    self.disable_input_fields()
                    self.timer.start(1000)
                    
    def disable_input_fields(self):
        self.lineEdit.setDisabled(True)
        self.lineEdit_2.setDisabled(True)
        self.pushButton.setDisabled(True)
        self.label_2.setText("30초 후에 가능합니다.")
        self.label_3.setHidden(True)
        
    def enable_input_fields(self):
        self.label_2.setText("다시 입력하세요.")
        self.lineEdit.setEnabled(True)
        self.lineEdit_2.setEnabled(True)
        self.pushButton.setEnabled(True)
        
    def exit_application(self):
        self.remaining_seconds -= 1
        
        if self.remaining_seconds == 0:
            self.timer.stop()
            self.enable_input_fields()
            self.incorrect_attempts = 0
            self.remaining_seconds = 5
            self.label_3.setHidden(True)
            
    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)                    
    MainWindow.show()
    sys.exit(app.exec_())
